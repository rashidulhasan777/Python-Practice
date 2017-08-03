#!python3
# Email Scrape from Oberlo

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import selenium.webdriver.support.ui as ui
import openpyxl, re, time

emailRegex=re.compile(r'([a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+(\.[a-zA-Z]{2,5}))')
countryRegex=re.compile(r'"shipping_country":".*?,')
nameRegex=re.compile(r'"shipping_name":".*?,')
orderRegex=re.compile(r'"order_name":".*?,')
wb=openpyxl.load_workbook('Email from Oberlo.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')

browser=webdriver.Chrome()
browser.get('https://bradleys-mens-store.myshopify.com/admin/apps')
emailElem=browser.find_element_by_id('login-input')
emailElem.send_keys('businesswithvideos@gmail.com')
passwordElem=browser.find_element_by_id('password')
passwordElem.send_keys('shosi12341')
passwordElem.send_keys(Keys.ENTER)
directoriesElem=browser.find_element_by_link_text('Oberlo')
directoriesElem.click()
url='https://app.oberlo.com/orders?from=2015-12-01&to=2017-04-08&page='
pageNo=3406
row=2
time.sleep(30)
while(pageNo != 100):
    browser.get(url+str(pageNo))
    source=browser.page_source
    moName=nameRegex.findall(source)
    moCountry=countryRegex.findall(source)
    moEmail=emailRegex.findall(source)
    moOrder=orderRegex.findall(source)
    for i in range(len(moName)):
        sheet.cell(row=row, column=1).value=moName[i].replace('"shipping_name":"', '').replace('",', '')
        sheet.cell(row=row, column=2).value=moEmail[i+31][0]
        sheet.cell(row=row, column=3).value=moCountry[i].replace('"shipping_country":"', '').replace('",', '')
        sheet.cell(row=row, column=4).value=moOrder[i].replace('"order_name":"', '').replace('",', '')
        row+=1
    wb.save('Email from Oberlo.xlsx')
    print(pageNo)
    pageNo-=1
print('Done')    
    
