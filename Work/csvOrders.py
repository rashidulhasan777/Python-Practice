import openpyxl

wbOne = openpyxl.load_workbook('orderInformation.xlsx')
sheetOne = wbOne.get_sheet_by_name('Sheet1')
wbTwo = openpyxl.load_workbook('orders_export.xlsx')
sheetTwo = wbTwo.get_sheet_by_name('orders_export')

sheetOneRow = 2

numberOfRowInSheetTwo = sheetTwo.max_row + 1
fromOrder = 3352

for sheetTwoRow in range(2, numberOfRowInSheetTwo):
    currentOrder = int(sheetTwo.cell(row=sheetTwoRow, column=1).value.replace('#', ''))
    if currentOrder < fromOrder:
        break
    if sheetTwo.cell(row=sheetTwoRow, column=3).value == 'refunded':
        continue
    if sheetTwo.cell(row=sheetTwoRow, column=1).value == sheetTwo.cell(row=sheetTwoRow-1, column=1).value:
        sheetOne.cell(row=sheetOneRow, column=1).value = 'ePacket'
        sheetOne.cell(row=sheetOneRow, column=2).value = sheetTwo.cell(row=sheetTwoRow, column=17).value
        sheetOne.cell(row=sheetOneRow, column=3).value = sheetOne.cell(row=sheetOneRow-1, column=3).value
        sheetOne.cell(row=sheetOneRow, column=4).value = sheetTwo.cell(row=sheetTwoRow, column=18).value
        sheetOne.cell(row=sheetOneRow, column=5).value = ''
        sheetOne.cell(row=sheetOneRow, column=6).value = sheetOne.cell(row=sheetOneRow-1, column=6).value
        sheetOne.cell(row=sheetOneRow, column=7).value = sheetOne.cell(row=sheetOneRow-1, column=7).value
        sheetOne.cell(row=sheetOneRow, column=8).value = sheetOne.cell(row=sheetOneRow-1, column=8).value
        sheetOne.cell(row=sheetOneRow, column=9).value = sheetOne.cell(row=sheetOneRow-1, column=9).value
        sheetOne.cell(row=sheetOneRow, column=10).value = sheetOne.cell(row=sheetOneRow-1, column=10).value
        sheetOne.cell(row=sheetOneRow, column=11).value = sheetOne.cell(row=sheetOneRow-1, column=11).value
        sheetOne.cell(row=sheetOneRow, column=12).value = sheetOne.cell(row=sheetOneRow-1, column=12).value
        sheetOne.cell(row=sheetOneRow, column=13).value = sheetOne.cell(row=sheetOneRow-1, column=13).value
        sheetOne.cell(row=sheetOneRow, column=14).value = sheetOne.cell(row=sheetOneRow-1, column=14).value
        sheetOne.cell(row=sheetOneRow, column=15).value = ''
        sheetOne.cell(row=sheetOneRow, column=16).value = sheetOne.cell(row=sheetOneRow-1, column=16).value
        sheetOne.cell(row=sheetOneRow, column=17).value = sheetOne.cell(row=sheetOneRow-1, column=17).value
        sheetOneRow += 1
        wbOne.save('orderInformation.xlsx')
        continue
    if sheetTwo.cell(row=sheetTwoRow, column=38).value:
        addressLine2 = str(sheetTwo.cell(row=sheetTwoRow, column=38).value) + ', '
    else:
        addressLine2=''
    fullAddress = sheetTwo.cell(row=sheetTwoRow, column=37).value + ', ' + addressLine2
    try:
        fullAddress += sheetTwo.cell(row=sheetTwoRow, column=40).value + ', ' + sheetTwo.cell(row=sheetTwoRow, column=42).value + ', '
        fullAddress += sheetTwo.cell(row=sheetTwoRow, column=43).value
    except:
        if not sheetTwo.cell(row=sheetTwoRow, column=40).value:
            x = ''
        else:
            x = sheetTwo.cell(row=sheetTwoRow, column=40).value + ', '
        if not sheetTwo.cell(row=sheetTwoRow, column=42).value:
            y =  ''
        else:
            y = sheetTwo.cell(row=sheetTwoRow, column=42).value + ', '
        if not sheetTwo.cell(row=sheetTwoRow, column=43).value:
            z = ''
        else:
            z = sheetTwo.cell(row=sheetTwoRow, column=43).value
        fullAddress += x + y + z



    sheetOne.cell(row=sheetOneRow, column=1).value = 'ePacket'
    sheetOne.cell(row=sheetOneRow, column=2).value = sheetTwo.cell(row=sheetTwoRow, column=17).value
    sheetOne.cell(row=sheetOneRow, column=3).value = sheetTwo.cell(row=sheetTwoRow, column=21).value
    sheetOne.cell(row=sheetOneRow, column=4).value = sheetTwo.cell(row=sheetTwoRow, column=18).value
    sheetOne.cell(row=sheetOneRow, column=5).value = ''
    sheetOne.cell(row=sheetOneRow, column=6).value = sheetTwo.cell(row=sheetTwoRow, column=25).value
    sheetOne.cell(row=sheetOneRow, column=7).value = sheetTwo.cell(row=sheetTwoRow, column=2).value
    sheetOne.cell(row=sheetOneRow, column=8).value = fullAddress
    sheetOne.cell(row=sheetOneRow, column=9).value = sheetTwo.cell(row=sheetTwoRow, column=35).value
    sheetOne.cell(row=sheetOneRow, column=10).value = sheetTwo.cell(row=sheetTwoRow, column=43).value
    sheetOne.cell(row=sheetOneRow, column=11).value = sheetTwo.cell(row=sheetTwoRow, column=42).value
    sheetOne.cell(row=sheetOneRow, column=12).value = sheetTwo.cell(row=sheetTwoRow, column=40).value
    sheetOne.cell(row=sheetOneRow, column=13).value = sheetTwo.cell(row=sheetTwoRow, column=37).value + ', ' + addressLine2
    sheetOne.cell(row=sheetOneRow, column=14).value = sheetTwo.cell(row=sheetTwoRow, column=41).value
    sheetOne.cell(row=sheetOneRow, column=15).value = ''
    sheetOne.cell(row=sheetOneRow, column=16).value = sheetTwo.cell(row=sheetTwoRow, column=67).value
    sheetOne.cell(row=sheetOneRow, column=17).value = sheetTwo.cell(row=sheetTwoRow, column=1).value
    sheetOneRow += 1
    wbOne.save('orderInformation.xlsx')
