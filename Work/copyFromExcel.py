import openpyxl

wb1=openpyxl.load_workbook('PoolData.xlsx')
sheet1=wb1.get_sheet_by_name('Pool Data')
wb2=openpyxl.load_workbook('imortantPoolData.xlsx')
sheet2=wb2.get_sheet_by_name('Sheet1')
row2=2

for row1 in range(2, 64749):
    description=sheet1.cell(row=row1, column=8).value
    try:
        if 'pool' in description.lower():
            date=str(sheet1.cell(row=row1, column=3).value)
            if int(date[:4])==2016 and int(date[4:6])>=7:
                sheet2.cell(row=row2, column=1).value=sheet1.cell(row=row1, column=6).value
                sheet2.cell(row=row2, column=2).value=date[4:6]+'/'+date[6:]+'/'+date[:4]
                sheet2.cell(row=row2, column=3).value=sheet1.cell(row=row1, column=1).value
                sheet2.cell(row=row2, column=4).value=sheet1.cell(row=row1, column=4).value
                sheet2.cell(row=row2, column=5).value=sheet1.cell(row=row1, column=5).value
                sheet2.cell(row=row2, column=6).value=sheet1.cell(row=row1, column=8).value
                row2+=1
    except:
        continue

wb2.save('imortantPoolData.xlsx')
print('Done')
