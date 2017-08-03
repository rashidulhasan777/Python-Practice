import openpyxl
wb=openpyxl.load_workbook('Email from Oberlo.xlsx')
sheet=wb.get_sheet_by_name('Sheet1')

for i in range(2, 17537):
    if sheet.cell(row=i, column=3).value=='United States':
        file=open('email.txt', 'a')
        file.write(sheet.cell(row=i, column=2).value+'\n')
        file.close()
