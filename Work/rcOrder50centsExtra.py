#! python3
import openpyxl

wb=openpyxl.load_workbook('RC Orders .xlsx')
sheet=wb.get_sheet_by_name('Pricing Section')

for row in range(2, 488):
    previousValue=sheet.cell(row=row, column=6).value
    currentValue=sheet.cell(row=row, column=7).value
    try:
        if float(currentValue)-float(previousValue)>0.5:
            sheet.cell(row=row, column=9).value='Increased more than 50 cents'
    except:
        print('hoinai')
wb.save('RC Orders .xlsx')
        
